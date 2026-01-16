# aut_excel/plan23.py
# -*- coding: utf-8 -*-

import os
import re
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
import unicodedata
from datetime import datetime
import uuid

# --------- helpers planilha ---------
def autosize_columns(ws, df: pd.DataFrame, max_width=80, pad=2):
    for idx, col in enumerate(df.columns, start=1):
        series = df[col].astype(str).fillna("")
        max_len = max(len(str(col)), *(len(v) for v in series)) + pad
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len, max_width)

def apply_font(ws, name="Helvetica", size=8):
    f = Font(name=name, size=size)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                            min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = f

# --------- helpers parsing ---------
def norm(s):
    return re.sub(r"\s+", " ", str(s)).strip()

def norm_upper_nospace(s):
    return re.sub(r"\s+", "", str(s)).strip().upper()

def strip_accents(s: str) -> str:
    return "".join(ch for ch in unicodedata.normalize("NFKD", str(s)) if not unicodedata.combining(ch))

def find_header_row(df_raw):
    for i in range(len(df_raw)):
        row = df_raw.iloc[i]
        if any(norm_upper_nospace(x) == "FONTE" for x in row):
            return i
    return None

def re_find_col(columns, target="FONTE"):
    for c in columns:
        if norm_upper_nospace(c) == norm_upper_nospace(target):
            return c
    for c in columns:
        if norm_upper_nospace(target) in norm_upper_nospace(c):
            return c
    return None

def only_8_digits(val):
    if pd.isna(val): return None
    s = str(val).strip()
    m = re.search(r"(?<!\d)(\d{8})(?!\d)", s)
    if m: return m.group(1)
    try:
        i = int(float(s))
        s2 = f"{i:d}"
        if len(s2) == 8 and s2.isdigit():
            return s2
    except Exception:
        pass
    return None

RX_GRUPO  = re.compile(r"^\s*\d\s*-\s+.+", re.IGNORECASE)
RX_QUADRO = re.compile(r"^\s*[a-fA-F]\.\s+.+")
RX_TOTAL  = re.compile(r"^\s*Total\s+da\s+Fonte\s*:?\s*$", re.IGNORECASE)

def br_to_float(x):
    if x is None or (isinstance(x, float) and pd.isna(x)): return None
    s = str(x).strip()
    if s == "": return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None

# --------- mapeamentos ---------
FONTE_MAP = {
    "15000000": "15000000 - Recursos não vinculados de Impostos",
    "15001001": "15001001 - Recursos destinados à Manutenção e Desenvolvimento do Ensino",
    "15010000": "15010000 - Outros Recursos não Vinculados",
    "15010100": "15010100 - Outros Recursos não vinculados destinados ao Tesouro",
    "15400000": "15400000 - Transferência de recursos do FUNDEB desenvolvimento do Ensino",
    "15401070": "15401070 - Transferência de recursos do FUNDEB Remuneração Educação Básica",
    "15500000": "15500000 - Recursos da Contribuição ao Salário Educação",
    "15510000": "15510000 - Transferências de Recursos do FNDE referente ao Programa Dinheiro Direto na Escola (PDDE)",
    "15520000": "15520000 - Transferências de Recursos do FNDE referente ao Programa Nacional de Alimentação Escolar (PNAE)",
    "15530000": "15530000 - Transferências de Recursos do FNDE referente ao P. N. de Apoio ao Transporte Escolar (PNATE)",
    "15690000": "15690000 - Outras Transferências de Recursos do FNDE",
    "15700000": "15700000 - Transferências do Governo Federal ref. a Convênios e outros Repasses vinculados à Educação",
    "15740000": "15740000 - Recursos de Operações de Crédito Educação",
}

def remap_ci(series: pd.Series, mapping: dict) -> pd.Series:
    norm_map = {norm(strip_accents(k)).upper(): v for k, v in mapping.items()}
    def _one(x):
        key = norm(strip_accents(x)).upper()
        return norm_map.get(key, x)
    return series.fillna("").map(_one)

GRUPO_MAP = {
    "3 - OUTRAS DESPESAS CORRENTES": "3 - Outras Despesas Corrente",
    "4 - INVESTIMENTOS": "4 - Investimentos",
    "1 - PESSOAL E ENCARGOS SOCIAIS": "1 - Pessoal e Encargos Sociais",
}

SUBTETO_MAP = {
    "a. Despesas Obrigatórias": "A - Despesas Obrigatórias",
    "b. Essenciais à Manutenção da Unidade": "B - Essenciais à Manutenção da Unidade",
    "c. Despesas Prioridades Estratégicas": "C - Prioridades Estratégicas LDO",
    "d. Despesas Essenciais Finalísticas": "D - Essenciais Finalísticas",
    "e. Projetos de Investimentos": "E - Projetos de Investimentos",
    "f. Demais ações e projetos": "F - Demais Ações e Projetos Finalísticos",
}

# --------- Exercício no cabeçalho ---------
def extract_exercicio(df_raw) -> str | None:
    rx = re.compile(r"exerc[ií]cio\s+igual\s+a\s+(\d{4})", re.IGNORECASE)
    max_rows = min(50, len(df_raw))
    for i in range(max_rows):
        for val in df_raw.iloc[i].tolist():
            txt = strip_accents(str(val))
            m = rx.search(txt)
            if m:
                return m.group(1)
    return None

# --------- função chamável ---------
def gerar_plan23_tratado(input_path: str, output_dir: str) -> str:
    """
    Lê o Excel 'input_path' (relatório Plan 23), gera um arquivo com as abas:
      - plan23_tratado
      - plan23_final
      - plan23_bd
    Salvando em 'output_dir' com nome único. Retorna o caminho completo.
    """
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {input_path}")
    os.makedirs(output_dir, exist_ok=True)
    unique = datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:8]
    output_path = os.path.join(output_dir, f"plan23_tratado-{unique}.xlsx")

    # 1) ler bruta
    df_raw = pd.read_excel(input_path, sheet_name=0, header=None, dtype=str)
    exercicio = extract_exercicio(df_raw)

    hdr = find_header_row(df_raw)
    if hdr is None:
        raise RuntimeError("Não encontrei a linha de cabeçalho com 'FONTE'.")

    cols = df_raw.iloc[hdr].tolist()
    df = df_raw.iloc[hdr+1:].copy()
    df.columns = cols

    col_fonte = re_find_col(df.columns, "FONTE")
    col_gq    = re_find_col(df.columns, "GRUPO DE DESPESA / QUADRO ORÇAMENTÁRIO")
    col_teto  = re_find_col(df.columns, "TETO ANUAL")
    col_saldo = re_find_col(df.columns, "SALDO ANUAL")
    if not all([col_fonte, col_gq, col_teto, col_saldo]):
        raise RuntimeError("Cabeçalhos necessários não encontrados (FONTE, GRUPO..., TETO ANUAL, SALDO ANUAL).")

    registros = []
    fonte_atual = None
    grupo_atual = None

    for _, row in df.iterrows():
        fonte_cell = row.get(col_fonte, "")
        gq_cell    = norm(row.get(col_gq, ""))
        teto_cell  = norm(row.get(col_teto, ""))
        saldo_cell = norm(row.get(col_saldo, ""))

        cod8 = only_8_digits(fonte_cell)
        if cod8:
            fonte_atual = cod8
            grupo_atual = gq_cell if RX_GRUPO.match(gq_cell or "") else None
            continue

        if not fonte_atual:
            continue

        if RX_TOTAL.match(gq_cell or ""):
            fonte_atual = None
            grupo_atual = None
            continue

        if RX_QUADRO.match(gq_cell or "") and grupo_atual:
            registros.append({
                "Fonte": fonte_atual,
                "Grupo de Despesa": grupo_atual,
                "Quadro Orçamentário": gq_cell,
                "Teto Anual": teto_cell,
                "Saldo Anual": saldo_cell,
            })

    trat = pd.DataFrame(registros)

    # plan23_final
    trat["_teto_num"] = trat["Teto Anual"].apply(br_to_float)
    final = trat[trat["_teto_num"].fillna(0) != 0].copy()
    final = final.rename(columns={
        "Fonte": "fonte",
        "Grupo de Despesa": "grupo_despesa",
        "Quadro Orçamentário": "subteto_despesa_momp",
        "Teto Anual": "teto_anual",
    })[["fonte", "grupo_despesa", "subteto_despesa_momp", "teto_anual"]]

    # padronizações
    final["fonte"] = final["fonte"].astype(str).str.strip().map(lambda x: FONTE_MAP.get(x, x))
    final["grupo_despesa"] = remap_ci(final["grupo_despesa"], GRUPO_MAP)
    final["subteto_despesa_momp"] = remap_ci(final["subteto_despesa_momp"], SUBTETO_MAP)

    # plan23_bd
    bd = final.copy()
    gd_idx = bd.columns.get_loc("grupo_despesa")
    bd.insert(gd_idx + 1, "teto_despesa_momp", "4 - A Classificar")
    bd.insert(0, "exercicio", exercicio or "")

    # grava
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        trat.drop(columns=["_teto_num"], errors="ignore").to_excel(writer, sheet_name="plan23_tratado", index=False)
        ws1 = writer.sheets["plan23_tratado"]
        autosize_columns(ws1, trat.drop(columns=["_teto_num"], errors="ignore"))
        apply_font(ws1, "Helvetica", 8)

        final.to_excel(writer, sheet_name="plan23_final", index=False)
        ws2 = writer.sheets["plan23_final"]
        autosize_columns(ws2, final)
        apply_font(ws2, "Helvetica", 8)

        bd.to_excel(writer, sheet_name="plan23_bd", index=False)
        ws3 = writer.sheets["plan23_bd"]
        autosize_columns(ws3, bd)
        apply_font(ws3, "Helvetica", 8)

    return output_path

# Execução direta (opcional para testes locais)
if __name__ == "__main__":
    raise SystemExit("Use gerar_plan23_tratado(input_path, output_dir) a partir do sistema.")
