# qompxpta.py
# -*- coding: utf-8 -*-

import os
import re
import unicodedata
import uuid
from datetime import datetime
import pandas as pd
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

# Colunas (Plan 23)
P_AOEE   = "Ação (PAOE)"
P_PROD   = "Produto da Ação"
P_SUB    = "Subação/entrega"
P_FONTE  = "Fonte"
P_GRUPO  = "Grupo"
P_SUBT   = "Tipificação da Despesa"
P_VALOR  = "Valor PTA"

def br_to_float(x):
    if pd.isna(x):
        return None
    s = str(x).strip()
    if s == "":
        return None
    s = s.replace(".", "").replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None

def float_to_brl(x, ndigits=2):
    if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and x.strip()==""):
        return ""
    try:
        v = float(str(x).replace(".", "").replace(",", ".")) if isinstance(x, str) else float(x)
    except Exception:
        return str(x)
    s = f"{v:,.{ndigits}f}"
    s = s.replace(",", "X").replace(".", ",").replace("X", ".")
    return s

def float_to_us_str(x, ndigits=2):
    """
    Converte para string US sem destruir decimais:
    - Se houver vírgula, assume BR -> remove '.' e troca ',' por '.'
    - Se não houver vírgula, assume US -> mantém '.' como decimal
    """
    if x is None or (isinstance(x, float) and pd.isna(x)) or (isinstance(x, str) and x.strip()==""):
        return ""
    s = str(x).strip()
    if "," in s:
        # formato BR (p.ex. 38.049,00)
        s = s.replace(".", "").replace(",", ".")
    # se não tem vírgula, já está em US; não remover o ponto!
    try:
        v = float(s)
    except Exception:
        return ""
    return f"{v:.2f}"

def strip_accents(s: str) -> str:
    s = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in s if not unicodedata.combining(ch))

def super_norm(s) -> str:
    if pd.isna(s):
        return ""
    s = str(s).strip().lower()
    s = strip_accents(s)
    s = s.replace("-", " ").replace("/", " ").replace("\\", " ")
    s = re.sub(r"[^\w\s]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()

def split_subacao(valor: str):
    if pd.isna(valor):
        return "", ""
    s = str(valor)
    first = s.find("*")
    last  = s.rfind("*")
    if first == -1 and last == -1:
        return "", s.strip()
    if first != -1 and last != -1 and last > first:
        chave = s[first:last+1].strip()
        desc  = s[last+1:].strip()
        return chave, desc
    if first != -1:
        return s[first:].strip(), s[:first].strip()
    if last != -1:
        return s[:last+1].strip(), s[last+1:].strip()
    return "", s.strip()

def parse_chave_oito(chave: str):
    if pd.isna(chave):
        return ("", "", "", "", "", "", "", "")
    parts = [p.strip() for p in str(chave).split("*")]
    parts = [p for p in parts if p != ""]
    if len(parts) < 8:
        parts += [""] * (8 - len(parts))
    return tuple(parts[:8])

def autosize_columns(worksheet, df: pd.DataFrame, max_width=80, pad=2):
    for idx, col in enumerate(df.columns, 1):
        series = df[col].astype(str).replace("None", "")
        max_len = max([len(str(col))] + [len(s) for s in series.tolist()]) + pad
        max_len = min(max_len, max_width)
        worksheet.column_dimensions[get_column_letter(idx)].width = max_len

def apply_font(ws, font_name="Helvetica", font_size=8):
    f = Font(name=font_name, size=font_size)
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.font = f

def remap_ci_accents(series: pd.Series, mapping: dict) -> pd.Series:
    def strip_acc(s: str) -> str:
        return "".join(ch for ch in unicodedata.normalize("NFKD", s) if not unicodedata.combining(ch))
    upper_map = {strip_acc(k).upper().strip(): v for k, v in mapping.items()}
    def _one(x):
        key = strip_acc(str(x)).upper().strip()
        return upper_map.get(key, x)
    return series.apply(_one)

GRUPO_MAP = {
    "4-INVESTIMENTOS": "4 - Investimentos",
    "3-OUTRAS DESPESAS CORRENTES": "3 - Outras Despesas Corrente",
    "1-PESSOAL E ENCARGOS SOCIAIS": "1 - Pessoal e Encargos Sociais",
}
SUBTETO_MAP = {
    "Demais Ações e Projetos": "F - Demais Ações e Projetos Finalísticos",
    "Despesas Essenciais Finalísticas": "D - Essenciais Finalísticas",
    "Despesas Obrigatórias": "A - Despesas Obrigatórias",
    "Despesas Prioridades Estratégicas": "C - Prioridades Estratégicas LDO",
    "Essenciais à Manutenção da Unidade": "B - Essenciais à Manutenção da Unidade",
}
ACAO_MAP = {
    "2009 -Manutenção de ações de informática": "2009 - Manutenção de ações de informática",
    "2010 -Manutenção de órgãos colegiados": "2010 - Manutenção de órgãos colegiados",
    "2014 -Publicidade institucional e propaganda": "2014 - Publicidade institucional e propaganda",
    "2284 -Manutenção do Conselho Estadual de Educação - CEE": "2284 - Manutenção do Conselho Estadual de Educação - CEE",
    "2895 -Alimentação Escolar da Educação de Jovens e Adultos": "2895 - Alimentação Escolar da Educação de Jovens e Adultos",
    "2897 -Alimentação Escolar da Educação Especial": "2897 - Alimentação Escolar da Educação Especial",
    "2898 -Alimentação Escolar do Ensino Fundamental": "2898 - Alimentação Escolar do Ensino Fundamental",
    "2899 -Alimentação Escolar do Ensino Médio": "2899 - Alimentação Escolar do Ensino Médio",
    "2900 -Desenvolvimento da Educação de Jovens e Adultos": "2900 - Desenvolvimento da Educação de Jovens e Adultos",
    "2936 -Desenvolvimento das Modalidades de Ensino": "2936 - Desenvolvimento das Modalidades de Ensino",
    "2957 -Desenvolvimento da Educação Especial": "2957 - Desenvolvimento da Educação Especial",
    "4172 -Desenvolvimento do Ensino Fundamental": "4172 - Desenvolvimento do Ensino Fundamental",
    "4173 -Infraestrutura do Ensino Fundamental": "4173 - Infraestrutura do Ensino Fundamental",
    "4174 -Desenvolvimento do Ensino Médio": "4174 - Desenvolvimento do Ensino Médio",
    "4175 -Infraestrutura da Educação de Jovens e Adultos": "4175 - Infraestrutura da Educação de Jovens e Adultos",
    "4177 -Infraestrutura do Ensino Médio": "4177 - Infraestrutura do Ensino Médio",
    "4178 -Infraestrutura da Educação Especial": "4178 - Infraestrutura da Educação Especial",
    "4179 -Transporte Escolar da Educação Especial": "4179 - Transporte Escolar da Educação Especial",
    "4180 -Infraestrutura de Administração e Gestão": "4180 - Infraestrutura de Administração e Gestão",
    "4181 -Transporte Escolar do Ensino Fundamental": "4181 - Transporte Escolar do Ensino Fundamental",
    "4182 -Transporte Escolar do Ensino Médio": "4182 - Transporte Escolar do Ensino Médio",
    "4491 -Pagamento de verbas indenizatórias a servidores estaduais.": "4491 - Pagamento de verbas indenizatórias a servidores estaduais",
    "4524 -FMTE - Ensino Fundamental": "4524 - FMTE - Ensino Fundamental",
    "4525 -FMTE - Educação Infantil": "4525 - FMTE - Educação Infantil",
    "8002 -Recolhimento do PIS-PASEP e pagamento do abono": "8002 - Recolhimento do PIS-PASEP e pagamento do abono",
    "8003 -Cumprimento de sentenças judiciais transitadas em julgado - Adm. Direta": "8003 - Cumprimento de sentenças judiciais transitadas em julgado - Adm. Direta",
    "8040 -Recolhimento de encargos e obrigações previdenciárias de inativos e pensionistas do Estado de Mato Grosso": "8040 - Recolhimento de encargos e obrigações previdenciárias de inativos e pensionistas do Estado de Mato Grosso",
}

def gerar_qompxpta_tratado(input_path: str, output_dir: str) -> str:
    """
    Lê o Excel 'input_path' (primeira aba), aplica as regras e grava um arquivo
    tratado (NOME ÚNICO) em 'output_dir'. Retorna o caminho completo do arquivo.
    """
    if not os.path.isfile(input_path):
        raise FileNotFoundError(f"Arquivo não encontrado: {input_path}")

    os.makedirs(output_dir, exist_ok=True)
    unique = datetime.now().strftime("%Y%m%d-%H%M%S") + "-" + uuid.uuid4().hex[:8]
    out_path = os.path.join(output_dir, f"qompxpta_tratado-{unique}.xlsx")

    # 1) Ler arquivo (1ª aba)
    plan_dict = pd.read_excel(input_path, sheet_name=None, dtype=str)
    df_plan = list(plan_dict.values())[0].copy()

    # 2) Conferir colunas obrigatórias do Plan 23
    req_plan = [P_AOEE, P_PROD, P_SUB, P_FONTE, P_GRUPO, P_SUBT, P_VALOR]
    miss_p = [c for c in req_plan if c not in df_plan.columns]
    if miss_p:
        raise ValueError(f"Colunas ausentes no arquivo: {miss_p}")

    # 3) Preparar Plan 23
    df_plan[P_VALOR] = df_plan[P_VALOR].apply(br_to_float)
    for col in [P_AOEE, P_PROD, P_SUB, P_FONTE, P_GRUPO, P_SUBT]:
        df_plan[col] = df_plan[col].astype(str).str.strip()

    # Agregar duplicados por 6 chaves somando Valor PTA
    grp_cols_plan = [P_AOEE, P_PROD, P_SUB, P_FONTE, P_GRUPO, P_SUBT]
    df_plan_agg = (
        df_plan.groupby(grp_cols_plan, dropna=False, as_index=False)[P_VALOR]
               .sum(min_count=1)
    )

    # 4) Split da Subação em chave + descrição
    chaves, descrs = [], []
    for v in df_plan_agg[P_SUB].fillna(""):
        chave, desc = split_subacao(v)
        chaves.append(chave)
        descrs.append(desc)
    df_plan_split = df_plan_agg.copy()
    df_plan_split["Chave de Planejamento"] = chaves
    df_plan_split["Descrição subação"]      = descrs

    # 5) plan23_tratada
    cols_tratada = ["Chave de Planejamento", P_AOEE, P_PROD, P_FONTE, P_GRUPO, P_SUBT, P_VALOR]
    df_plan_tratada = (
        df_plan_split[cols_tratada]
            .groupby(["Chave de Planejamento", P_AOEE, P_PROD, P_FONTE, P_GRUPO, P_SUBT],
                     dropna=False, as_index=False)[P_VALOR]
            .sum(min_count=1)
    )

    # 6) Explodir a chave em 8 partes e montar plan134_chave
    reg, subf, adj, macro, pilar, eixo, pol_dec_value, pub_trans = [], [], [], [], [], [], [], []
    for v in df_plan_tratada["Chave de Planejamento"].fillna(""):
        r, sf, a, m, p, e, pd_val, pt = parse_chave_oito(v)
        reg.append(r); subf.append(sf); adj.append(a); macro.append(m)
        pilar.append(p); eixo.append(e); pol_dec_value.append(pd_val); pub_trans.append(pt)

    df_plan_chave = df_plan_tratada.copy()
    df_plan_chave.insert(0, "Região", reg)
    df_plan_chave.insert(1, "Subfunção_UG", subf)
    df_plan_chave.insert(2, "ADJ", adj)
    df_plan_chave.insert(3, "Macropolitica", macro)
    df_plan_chave.insert(4, "Pilar", pilar)
    df_plan_chave.insert(5, "Eixo", eixo)
    df_plan_chave.insert(6, "Politica Decreto", pol_dec_value)
    df_plan_chave.insert(7, "Publico Transversal", pub_trans)
    df_plan_chave = df_plan_chave[
        ["Região","Subfunção_UG","ADJ","Macropolitica","Pilar","Eixo",
         "Politica Decreto","Publico Transversal","Chave de Planejamento",
         P_AOEE,P_FONTE,P_GRUPO,P_SUBT,P_VALOR]
    ].copy()

    # 7) Limpar/renomear
    df_plan_chave = df_plan_chave.replace({"nan": pd.NA, "NaN": pd.NA, "None": pd.NA})
    df_plan_chave = df_plan_chave.replace(r"^\s*$", pd.NA, regex=True)
    df_plan_chave = df_plan_chave.dropna(how="all")
    key_cols_blank = ["Região","Subfunção_UG","ADJ","Macropolitica","Pilar","Eixo",
                      "Politica Decreto","Publico Transversal","Chave de Planejamento"]
    df_plan_chave = df_plan_chave.dropna(subset=key_cols_blank, how="all")

    rename_map = {
        "Região": "regiao",
        "Subfunção_UG": "subfuncao_ug",
        "ADJ": "adj",
        "Macropolitica": "macropolitica",
        "Pilar": "pilar",
        "Eixo": "eixo",
        "Politica Decreto": "politica_decreto",
        "Publico Transversal": "publico_transversal",
        "Chave de Planejamento": "chave_planejamento",
        "Ação (PAOE)": "acao_paoe",
        "Fonte": "fonte",
        "Grupo": "grupo_despesa",
        "Tipificação da Despesa": "subteto_despesa_momp",
        "Valor PTA": "teto_politica_decreto",
    }
    df_plan_chave = df_plan_chave.rename(columns=rename_map)

    # mapeamentos
    df_plan_chave["grupo_despesa"] = remap_ci_accents(df_plan_chave["grupo_despesa"], GRUPO_MAP)
    df_plan_chave["subteto_despesa_momp"] = remap_ci_accents(df_plan_chave["subteto_despesa_momp"], SUBTETO_MAP)
    df_plan_chave["acao_paoe"] = remap_ci_accents(df_plan_chave["acao_paoe"], ACAO_MAP)

    # ordem final
    final_cols = [
        "regiao","subfuncao_ug","adj","macropolitica","pilar","eixo",
        "politica_decreto","publico_transversal","chave_planejamento",
        "acao_paoe","fonte","grupo_despesa","subteto_despesa_momp","teto_politica_decreto"
    ]
    df_plan_chave = df_plan_chave.reindex(columns=final_cols)

    # versões BR / US
    df_plan_chave_br = df_plan_chave.copy()
    df_plan_chave_br["teto_politica_decreto"] = df_plan_chave_br["teto_politica_decreto"].apply(float_to_brl)

    df_plan_final = df_plan_chave.copy()
    df_plan_final["teto_politica_decreto"] = df_plan_final["teto_politica_decreto"].apply(float_to_us_str)

    # 8) Exportar Excel (arquivo de saída com nome único)
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        # renomeadas:
        df_plan_chave_br.to_excel(writer, sheet_name="plan134_chave", index=False)
        ws_brl = writer.sheets["plan134_chave"]
        autosize_columns(ws_brl, df_plan_chave_br)
        apply_font(ws_brl, font_name="Helvetica", font_size=8)

        df_plan_final.to_excel(writer, sheet_name="plan134_final", index=False)
        ws_us = writer.sheets["plan134_final"]
        autosize_columns(ws_us, df_plan_final)
        apply_font(ws_us, font_name="Helvetica", font_size=8)

        writer.book.active = writer.book.sheetnames.index("plan134_chave")

    return out_path

# utilidades exportadas (usadas pelo módulo chamador)
__all__ = ["gerar_qompxpta_tratado", "autosize_columns", "apply_font"]
